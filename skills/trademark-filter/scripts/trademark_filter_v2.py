#!/usr/bin/env python3
"""
商标筛选工具 V2 (现代高级感审美版)
核心策略:
1. 反土味处罚：严厉打击微商感、年代感用字组合
2. 美学流派加分：奖励东方极简、院线科技、现代独立风
"""
import json
import re
import unicodedata
from collections import Counter
from typing import List, Dict

# ============================================================
# 美学词库配置 (Aesthetics Configuration)
# ============================================================

# 1. 土味/年代感警报词 (Tacky/Outdated) —— 严打
TACKY_CHARS = set('芬芳莲娇媚艳莎妮妃黛娜玛迪丹彩丝柔美美容丽娟萍霞翠红')
# 常见微商组合特征后缀 (过时)
TACKY_SUFFIXES = ['宝宝', '家族', '世家', '阿姨', '日记', '秘密']

# 2. 东方极简/禅意美学 (Oriental & Zen) —— 大幅加分
ORIENTAL_CHARS = set('观寻拾初本素朴归隐半未兮子言语溪谷木知山海风涧露茶墨白青玄一之若如')

# 3. 院线科技/极客护肤 (Clinical & Lab) —— 加分
LAB_CHARS = set('研因态极零玑科理医源秘赋微密码玻态皙修')

# 4. 现代独立/法式Chic (Modern & Chic) —— 加分
MODERN_CHARS = set('光悦己印纪彩幻星谜绘羽漾谧颜')

# 5. 常见低端成分字（需避雷，避免像廉价化妆品）
CHEAP_INGREDIENTS = set('参参茸茸蜜蜜膏膏油泥')

# 英文高质量词根 (现代极简风格)
ENGLISH_MODERN_ROOTS = [
    'AURA', 'LUMI', 'VERA', 'NOVA', 'DERMA', 'CLIN', 'TRUE', 'PURE',
    'NIVE', 'SENS', 'CELL', 'BIOM', 'GEN', 'BOTAN', 'HERB', 'ESSEN',
    'MINIMAL', 'BARE', 'RAW', 'NAKED', 'BLANC', 'SOLEIL', 'LUNE', 'TERRE',
    'EAU', 'AIR', 'ORCHID', 'OASIS', 'ZEN', 'SOUL', 'MIND'
]

# ============================================================
# 工具函数
# ============================================================

def is_chinese_char(c: str) -> bool:
    return '\u4e00' <= c <= '\u9fff'

def classify_language(name: str) -> str:
    if not name: return 'unknown'
    has_cn = any(is_chinese_char(c) for c in name)
    has_en = any(c.isascii() and c.isalpha() for c in name)
    if has_cn and has_en: return 'mixed'
    elif has_cn: return 'chinese'
    elif has_en: return 'english'
    return 'other'

def is_pro_english(name: str) -> bool:
    """更严格的英文可读性判断"""
    clean = name.replace(' ', '').upper()
    if len(clean) <= 3: return False
    
    # 元音检查
    vowels = set('AEIOUY') # 把Y当元音
    vowel_count = sum(1 for c in clean if c in vowels)
    if vowel_count == 0: return False
    if vowel_count / len(clean) > 0.6 or vowel_count / len(clean) < 0.2:
        return False
        
    consonants = set('BCDFGHJKLMNPQRSTVWXZ')
    # 连续辅音不得超过3个
    run = 0
    for c in clean:
        if c in consonants:
            run += 1
            if run >= 4: return False
        else:
            run = 0
    return True

# ============================================================
# V2 评分引擎
# ============================================================

def score_chinese_v2(name: str) -> Dict:
    """中文名深度审美评分 (满分100)"""
    scores = {
        '简洁高级感': 20,
        '意境与流派': 0,
        '音律美学': 15,
        '反土味惩罚': 0,
        '总分': 0
    }
    char_count = len(name)
    
    # 1. 简洁高级感 (最大 30)
    score_concise = 20
    if char_count == 2:
        score_concise = 30 # 两字品牌天然高级（如：逐本）
    elif char_count == 3:
        score_concise = 25
    elif char_count == 4:
        score_concise = 15
    else:
        score_concise = 0
    scores['简洁高级感'] = score_concise

    # 2. 意境与流派 (最大 40)
    score_vibe = 10
    
    oriental_hits = sum(1 for c in name if c in ORIENTAL_CHARS)
    lab_hits = sum(1 for c in name if c in LAB_CHARS)
    modern_hits = sum(1 for c in name if c in MODERN_CHARS)
    
    score_vibe += (oriental_hits * 10) # 东方极简权重最高
    score_vibe += (lab_hits * 8)
    score_vibe += (modern_hits * 7)
    
    # 如果完全没有命中任何高级词库，但字数少，给个保底
    if oriental_hits + lab_hits + modern_hits == 0 and char_count <= 3:
        score_vibe += 5
        
    scores['意境与流派'] = min(40, score_vibe)

    # 3. 音律美学 (最大 30)
    score_sound = 15
    # 叠词往往显得年轻或者有节奏（比如：观夏，不用叠词；但 橘朵 音律好）
    # 避免全部是同一部首的字（比如全是草字头，像草药名字）
    scores['音律美学'] = 20 if char_count <= 3 else 15

    # 4. 反土味惩罚 (负分机制！)
    penalty = 0
    tacky_hits = sum(1 for c in name if c in TACKY_CHARS)
    if tacky_hits >= 2:
        penalty -= 40 # 极重惩罚：如“娇丽”，“芬莎”
    elif tacky_hits == 1:
        penalty -= 15 # 警告性惩罚
        
    for suf in TACKY_SUFFIXES:
        if suf in name:
            penalty -= 50
            
    cheap_hits = sum(1 for c in name if c in CHEAP_INGREDIENTS)
    penalty -= (cheap_hits * 10)
    
    scores['反土味惩罚'] = penalty

    # 计算总分 (上限 100)
    scores['总分'] = max(0, min(100, score_concise + score_vibe + scores['音律美学'] + penalty))
    return scores

def score_english_v2(name: str) -> Dict:
    clean = name.strip()
    char_count = len(clean)
    scores = {'简洁高级感': 0, '意境与流派': 0, '音律美学': 0, '反土味惩罚': 0, '总分': 0}
    
    if not is_pro_english(name):
        return scores
        
    # 简洁高级感
    if char_count <= 5: scores['简洁高级感'] = 30
    elif char_count <= 8: scores['简洁高级感'] = 25
    elif char_count <= 10: scores['简洁高级感'] = 15
    else: scores['简洁高级感'] = 5
    
    # 意境
    upper = clean.upper()
    hits = sum(1 for root in ENGLISH_MODERN_ROOTS if root in upper)
    scores['意境与流派'] = min(40, 10 + hits * 15)
    
    scores['音律美学'] = 20
    scores['总分'] = min(100, sum(v for k, v in scores.items() if k != '总分'))
    return scores

# ============================================================
# 主筛选流程
# ============================================================

def filter_trademarks_v2(data: List[Dict], top_n: int = 50) -> List[Dict]:
    scored = []
    seen = set()
    
    for item in data:
        name = item.get('name', '').strip()
        if not name or name in seen:
            continue
        seen.add(name)
        
        lang = classify_language(name)
        
        # 抛弃纯数字等奇葩名字
        if re.search(r'\d', name):
            continue
            
        if lang == 'chinese':
            if len(name) > 4 or len(name) < 2: continue
            scores = score_chinese_v2(name)
        elif lang == 'english':
            if len(name) > 12 or len(name) < 4: continue
            scores = score_english_v2(name)
        elif lang == 'mixed':
            continue # V2 追求纯粹，放弃中英混搭（通常像山寨牌子）
        else:
            continue
            
        # 只要总分及格且惩罚项为0的
        if scores['总分'] >= 60 and scores['反土味惩罚'] == 0:
            scored.append({
                'name': name,
                'language': lang,
                'scores': scores,
                'total': scores['总分'],
                'original': item
            })
            
    # 打乱顺序，避免同分时总是按字母排序
    # 对于同分的，尽量选字数少的
    scored.sort(key=lambda x: (x['total'], -len(x['name'])), reverse=True)
    
    # 提取 TOP
    top_cn = [s for s in scored if s['language'] == 'chinese'][:40]
    top_en = [s for s in scored if s['language'] == 'english'][:10]
    
    final_list = top_cn + top_en
    final_list.sort(key=lambda x: x['total'], reverse=True)
    
    return final_list[:top_n]

def export_v2(results: List[Dict], output_path: str):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "大牌风·高级商标推荐"
    
    # 黑白极简设计风格 (V2美学)
    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='Microsoft YaHei', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    name_font = Font(name='Microsoft YaHei', size=14, bold=True, color='333333')
    
    thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
    
    headers = ['排名', '商标名', '总评', '意境流派', '基础简美', '降维惩罚', '图片URL']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for i, item in enumerate(results, 1):
        row = i + 1
        s = item['scores']
        
        ws.cell(row=row, column=1, value=i).alignment = data_alignment
        
        name_cell = ws.cell(row=row, column=2, value=item['name'])
        name_cell.font = name_font
        name_cell.alignment = data_alignment
        
        ws.cell(row=row, column=3, value=s['总分']).font = Font(bold=True)
        ws.cell(row=row, column=3).alignment = data_alignment
        
        ws.cell(row=row, column=4, value=s['意境与流派']).alignment = data_alignment
        ws.cell(row=row, column=5, value=s['简洁高级感'] + s['音律美学']).alignment = data_alignment
        ws.cell(row=row, column=6, value=s['反土味惩罚']).alignment = data_alignment
        ws.cell(row=row, column=7, value=item['original'].get('imgUrl', '')).font = Font(color='888888')
        
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border
            
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['G'].width = 60
    
    wb.save(output_path)
    print(f"✅ V2高级审美报告导出至: {output_path}")

def main():
    data_path = '/home/yangyy/projects/biaobiaoku-scraper/all_trademarks_0306_final.json'
    out_xls = '/home/yangyy/projects/biaobiaoku-scraper/antigravity-task/top50_trademarks_v2.xlsx'
    
    with open(data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    results = filter_trademarks_v2(data, 50)
    
    print("\n" + "="*50)
    print("✨ 极光/深空高级审美 —— V2 商标推荐单")
    print("="*50)
    for i, item in enumerate(results, 1):
        s = item['scores']
        # 加上星标强调极简名
        star = "★" if len(item['name']) == 2 else " "
        print(f"[{i:02d}] {star} {item['name']:<10s} | 总分:{s['总分']} | 意境:{s['意境与流派']}")
        
    export_v2(results, out_xls)

if __name__ == '__main__':
    main()
