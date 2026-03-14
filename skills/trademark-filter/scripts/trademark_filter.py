#!/usr/bin/env python3
"""
商标筛选工具 — 从日化商标数据中筛选 "简单、大气、好记" 的 TOP N 商标
Phase 3: 执行编码

评分体系 (满分100):
  - 简洁度 (25分): 字数 + 笔画复杂度
  - 音韵美感 (20分): 声调变化 + 开口度
  - 语义品质 (25分): 日化关联度 + 正面联想
  - 品牌感 (20分): 大气度 + 独特性
  - 记忆度 (10分): 画面感 + 语义完整性
"""
import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional

# ============================================================
# 配置区
# ============================================================

# 日化品类高品质用字（加分）
PREMIUM_CHARS = set('芙兰雪泉诗露韵妍馨莲萱黛碧珀瑶琳芬莎丽蔻美净润柔舒颜梦薇茉菲依伊欣怡姿')
# 自然/高端意象字
ELEGANT_CHARS = set('花月水云风竹梅荷桃樱松鹤玉金银珍珠')
# 清洁/护肤关联字
CLEAN_CHARS = set('洁清纯皙白亮晶透莹光')
# 大气/格局字
GRAND_CHARS = set('华盛皇御圣尊贵冠帝宝')
# 低品质/廉价感用字（减分）
LOW_QUALITY_CHARS = set('廉贱旧破烂臭腐')
# 负面含义字（减分）
NEGATIVE_CHARS = set('死鬼毒害凶恶丑败衰')

# 英文有意义的品牌词根
ENGLISH_GOOD_ROOTS = [
    'BLOOM', 'GLOW', 'PURE', 'GRACE', 'SILK', 'PEARL', 'ROSE', 'GENTLE',
    'RADIANT', 'BRIGHT', 'GOLDEN', 'SWEET', 'LOVE', 'BEAUTY', 'SOFT',
    'FRESH', 'FLORA', 'STAR', 'MOON', 'DREAM', 'ANGEL', 'PINK', 'CUTE',
    'HAPPY', 'JOY', 'DEAR', 'DARLING', 'PRETTY', 'LOVELY', 'ELEGANT',
    'CRYSTAL', 'DIAMOND', 'VELVET', 'ROYAL', 'NOBLE', 'CHARM', 'MAGIC',
    'PETITE', 'BELLE', 'BLISS', 'SERENE', 'DIVINE', 'LUXE', 'CHIC',
    'FLORAL', 'LILY', 'IRIS', 'VIOLET', 'ORCHID', 'LOTUS', 'JASMINE',
    'HONEY', 'COCO', 'MANGO', 'PEACH', 'BERRY', 'MINT', 'LEMON',
    'OCEAN', 'AQUA', 'SKY', 'CLOUD', 'SNOW', 'ICE', 'RAIN', 'DEW',
    'SWAN', 'RABBIT', 'DOVE', 'PANDA', 'KITTEN', 'TEDDY', 'MEOW',
    'GLAM', 'GLOW', 'SHINE', 'SPARK', 'GLEAM', 'SHIMMER', 'LUSTER',
    'CALM', 'ZEN', 'SPA', 'AURA', 'OASIS', 'NATURE', 'GREEN', 'ECO',
    'PERFECT', 'SUPER', 'WONDER', 'FANCY', 'FINE'
]

# 声母-韵母映射 (简化版)
# 开口度: 1=闭口, 2=半开, 3=开口
VOWEL_OPENNESS = {
    'a': 3, 'o': 3, 'e': 2, 'i': 1, 'u': 1, 'ü': 1,
    'ai': 3, 'ei': 2, 'ao': 3, 'ou': 2,
    'an': 3, 'en': 2, 'ang': 3, 'eng': 2,
    'ia': 3, 'ie': 2, 'iao': 3, 'iu': 1,
    'ian': 3, 'in': 1, 'iang': 3, 'ing': 1,
    'ua': 3, 'uo': 3, 'uai': 3, 'ui': 1,
    'uan': 3, 'un': 1, 'uang': 3, 'ong': 2,
    'üe': 2, 'üan': 3, 'ün': 1, 'iong': 2,
}

# ============================================================
# 工具函数
# ============================================================

def is_chinese_char(c: str) -> bool:
    """判断是否为中文字符"""
    return '\u4e00' <= c <= '\u9fff'

def is_common_char(c: str) -> bool:
    """判断是否为常用汉字（GB2312一级字表约3755字）"""
    # 简化判断：CJK统一汉字基本区内，且Unicode编码较小的通常更常见
    if not is_chinese_char(c):
        return False
    # 使用 unicodedata 检查字符是否有标准名称
    try:
        name = unicodedata.name(c, '')
        return bool(name)
    except:
        return True

def get_stroke_count(c: str) -> int:
    """估算汉字笔画数（基于Unicode CJK笔画数据的近似方法）"""
    if not is_chinese_char(c):
        return 0
    # 简化方法：使用字的复杂度作为代理
    # 常见简单字（笔画少）
    simple_chars = {
        '一': 1, '二': 2, '三': 3, '十': 2, '大': 3, '小': 3, '上': 3, '下': 3,
        '人': 2, '水': 4, '火': 4, '山': 3, '月': 4, '日': 4, '木': 4, '花': 7,
        '白': 5, '玉': 5, '兰': 5, '竹': 6, '云': 4, '风': 4, '雪': 11, '雨': 8,
        '美': 9, '丽': 7, '净': 8, '洁': 10, '清': 11, '纯': 7, '柔': 9, '润': 10,
        '芙': 7, '露': 20, '泉': 9, '诗': 8, '韵': 13, '妍': 7, '馨': 20, '莲': 10,
        '萱': 12, '黛': 17, '碧': 14, '珀': 10, '瑶': 14, '琳': 12, '芬': 7, '莎': 10,
        '蔻': 14, '舒': 12, '颜': 15, '梦': 11, '薇': 16, '茉': 8, '菲': 11,
        '依': 8, '伊': 6, '欣': 8, '怡': 9, '姿': 9, '皙': 13,
        '欧': 8, '法': 8, '韩': 12, '雅': 12, '植': 12,
    }
    if c in simple_chars:
        return simple_chars[c]
    # 默认估计
    code = ord(c)
    if code < 0x5000:
        return 8  # 常用字，偏简单
    elif code < 0x6000:
        return 10
    elif code < 0x7000:
        return 11
    elif code < 0x8000:
        return 12
    else:
        return 13

def classify_language(name: str) -> str:
    """分类商标语言类型"""
    if not name:
        return 'unknown'
    has_cn = any(is_chinese_char(c) for c in name)
    has_en = any(c.isascii() and c.isalpha() for c in name)
    if has_cn and has_en:
        return 'mixed'
    elif has_cn:
        return 'chinese'
    elif has_en:
        return 'english'
    else:
        return 'other'

def is_pronounceable_english(name: str) -> bool:
    """判断英文名是否可发音（非随机字母）"""
    clean = name.replace(' ', '').upper()
    if len(clean) <= 2:
        return False
    
    # 检查元音比例
    vowels = set('AEIOU')
    vowel_count = sum(1 for c in clean if c in vowels)
    if len(clean) > 0:
        vowel_ratio = vowel_count / len(clean)
        if vowel_ratio < 0.15:  # 元音太少，无法发音
            return False
        if vowel_ratio > 0.7:  # 元音太多，也不正常
            return False
    
    # 检查连续辅音长度
    consonant_run = 0
    max_consonant_run = 0
    for c in clean:
        if c not in vowels:
            consonant_run += 1
            max_consonant_run = max(max_consonant_run, consonant_run)
        else:
            consonant_run = 0
    
    if max_consonant_run >= 4:  # 4个连续辅音基本不可发音
        return False
    
    return True

def has_english_meaning(name: str) -> int:
    """检查英文名是否包含有意义的词根，返回匹配数"""
    upper = name.upper()
    count = 0
    for root in ENGLISH_GOOD_ROOTS:
        if root in upper:
            count += 1
    return count

# ============================================================
# 评分函数
# ============================================================

def score_chinese(name: str) -> Dict:
    """对中文商标名评分"""
    scores = {}
    
    # --- 维度1: 简洁度 (25分) ---
    char_count = len(name)
    if char_count == 2:
        length_score = 25
    elif char_count == 3:
        length_score = 20
    elif char_count == 4:
        length_score = 12
    else:
        length_score = max(0, 25 - char_count * 5)
    
    # 笔画复杂度调整
    avg_strokes = sum(get_stroke_count(c) for c in name) / max(len(name), 1)
    if avg_strokes <= 8:
        length_score = min(25, length_score + 3)
    elif avg_strokes >= 15:
        length_score = max(0, length_score - 5)
    elif avg_strokes >= 12:
        length_score = max(0, length_score - 2)
    
    scores['简洁度'] = min(25, length_score)
    
    # --- 维度2: 音韵美感 (20分) ---
    phonetic_score = 10  # 基础分
    
    # 声调变化奖励（简化：不同字的声调尽量不同）
    # 用字符编码的粗略分布来估计
    if char_count >= 2:
        # 检查是否有重复字（叠词有节奏感加分）
        if len(set(name)) < len(name):
            phonetic_score += 3  # 叠词加分
        # 字形多样性（不同部首）
        phonetic_score += min(5, len(set(name)) * 2)
    
    # 开口音奖励
    open_vowel_chars = set('芳花兰霞芙莎华达拉萨嘉')
    open_count = sum(1 for c in name if c in open_vowel_chars)
    phonetic_score += min(3, open_count * 2)
    
    scores['音韵美感'] = min(20, phonetic_score)
    
    # --- 维度3: 语义品质 (25分) ---
    semantic_score = 10  # 基础分
    
    # 日化关联字加分
    premium_count = sum(1 for c in name if c in PREMIUM_CHARS)
    semantic_score += min(8, premium_count * 3)
    
    # 自然/高端意象加分
    elegant_count = sum(1 for c in name if c in ELEGANT_CHARS)
    semantic_score += min(4, elegant_count * 2)
    
    # 清洁/护肤关联加分
    clean_count = sum(1 for c in name if c in CLEAN_CHARS)
    semantic_score += min(4, clean_count * 2)
    
    # 负面含义减分
    neg_count = sum(1 for c in name if c in NEGATIVE_CHARS)
    semantic_score -= neg_count * 5
    
    # 低品质感减分
    low_count = sum(1 for c in name if c in LOW_QUALITY_CHARS)
    semantic_score -= low_count * 3
    
    scores['语义品质'] = max(0, min(25, semantic_score))
    
    # --- 维度4: 品牌感/大气度 (20分) ---
    brand_score = 10  # 基础分
    
    # 大气字加分
    grand_count = sum(1 for c in name if c in GRAND_CHARS)
    brand_score += min(5, grand_count * 3)
    
    # 国际感/外来感加分（如"欧""法""韩"开头）
    intl_prefixes = set('欧法韩意英美日澳德')
    if name[0] in intl_prefixes:
        brand_score += 3
    
    # 字数2-3的天然品牌感更强
    if char_count == 2:
        brand_score += 4
    elif char_count == 3:
        brand_score += 2
    
    # 日化行业常见后缀加分
    good_suffixes_2 = ['泉', '诗', '兰', '露', '芙', '黛', '莎', '蔻']
    good_suffixes_3 = ['集', '堂', '坊', '阁']
    if char_count >= 2 and name[-1] in ''.join(good_suffixes_2):
        brand_score += 2
    
    scores['品牌感'] = min(20, brand_score)
    
    # --- 维度5: 记忆度 (10分) ---
    memory_score = 5  # 基础分
    
    # 语义完整性（有实际含义的字组合）
    if char_count <= 3:
        memory_score += 2
    
    # 叠词加分
    if len(name) >= 2 and len(set(name)) < len(name):
        memory_score += 2
    
    # 画面感（自然意象字）
    if any(c in ELEGANT_CHARS for c in name):
        memory_score += 1
    
    scores['记忆度'] = min(10, memory_score)
    
    # 总分
    scores['总分'] = sum(scores.values())
    
    return scores

def score_english(name: str) -> Dict:
    """对英文商标名评分"""
    scores = {}
    clean = name.strip()
    char_count = len(clean)
    
    # --- 可发音检查 ---
    if not is_pronounceable_english(clean):
        return {'简洁度': 0, '音韵美感': 0, '语义品质': 0, '品牌感': 0, '记忆度': 0, '总分': 0}
    
    # --- 维度1: 简洁度 (25分) ---
    if char_count <= 4:
        length_score = 25
    elif char_count <= 6:
        length_score = 22
    elif char_count <= 8:
        length_score = 18
    elif char_count <= 10:
        length_score = 14
    elif char_count <= 12:
        length_score = 10
    else:
        length_score = max(0, 25 - char_count)
    
    scores['简洁度'] = length_score
    
    # --- 维度2: 音韵美感 (20分) ---
    phonetic_score = 10
    # 有空格分词的更自然
    if ' ' in clean and len(clean.split()) == 2:
        phonetic_score += 3
    # 以常见品牌音节结尾
    nice_endings = ['A', 'E', 'O', 'Y', 'AN', 'IN', 'ON', 'ER', 'LE', 'LY', 'EL', 'AL']
    upper = clean.upper()
    for ending in nice_endings:
        if upper.endswith(ending):
            phonetic_score += 2
            break
    
    scores['音韵美感'] = min(20, phonetic_score)
    
    # --- 维度3: 语义品质 (25分) ---
    meaning_matches = has_english_meaning(clean)
    semantic_score = 5 + min(15, meaning_matches * 6)
    
    scores['语义品质'] = min(25, semantic_score)
    
    # --- 维度4: 品牌感 (20分) ---
    brand_score = 8
    if meaning_matches > 0:
        brand_score += min(8, meaning_matches * 4)
    if char_count <= 8:
        brand_score += 2
    # 首字母大写的正规感
    if clean[0].isupper():
        brand_score += 1
    
    scores['品牌感'] = min(20, brand_score)
    
    # --- 维度5: 记忆度 (10分) ---
    memory_score = 3
    if meaning_matches > 0:
        memory_score += 3
    if char_count <= 8:
        memory_score += 2
    if ' ' in clean:
        memory_score += 1  # 两词组合更有记忆点
    
    scores['记忆度'] = min(10, memory_score)
    
    scores['总分'] = sum(scores.values())
    return scores

def score_mixed(name: str) -> Dict:
    """对中英混合商标名评分"""
    # 提取中文部分单独评分，但品牌感加分
    cn_part = ''.join(c for c in name if is_chinese_char(c))
    en_part = ''.join(c for c in name if c.isascii() and c.isalpha())
    
    if cn_part:
        scores = score_chinese(cn_part)
        # 混合名天然更长，但如果中文部分简洁则补偿
        if len(cn_part) <= 3:
            scores['品牌感'] = min(20, scores.get('品牌感', 0) + 3)
        # 总长度惩罚
        total_len = len(name)
        if total_len > 8:
            scores['简洁度'] = max(0, scores['简洁度'] - 5)
            scores['记忆度'] = max(0, scores['记忆度'] - 2)
        scores['总分'] = scores['简洁度'] + scores['音韵美感'] + scores['语义品质'] + scores['品牌感'] + scores['记忆度']
        return scores
    else:
        return score_english(name)

# ============================================================
# 主筛选流程
# ============================================================

def filter_trademarks(data: List[Dict], top_n: int = 50) -> List[Dict]:
    """主筛选函数"""
    
    print(f"📊 开始筛选，共 {len(data)} 条商标\n")
    
    # === 第一轮：硬性过滤 ===
    print("=== 第一轮：硬性过滤 ===")
    filtered = []
    reject_reasons = Counter()
    
    for item in data:
        name = item.get('name', '').strip()
        if not name:
            reject_reasons['空名称'] += 1
            continue
        
        lang = classify_language(name)
        
        # 过滤条件
        if lang == 'chinese':
            if len(name) > 4:
                reject_reasons['中文超4字'] += 1
                continue
            if len(name) < 2:
                reject_reasons['中文少于2字'] += 1
                continue
            if any(c in NEGATIVE_CHARS for c in name):
                reject_reasons['含负面字'] += 1
                continue
        elif lang == 'english':
            clean = name.replace(' ', '')
            if len(clean) > 14:
                reject_reasons['英文超14字符'] += 1
                continue
            if len(clean) < 3:
                reject_reasons['英文少于3字符'] += 1
                continue
            if not is_pronounceable_english(name):
                reject_reasons['英文不可发音'] += 1
                continue
        elif lang == 'mixed':
            if len(name) > 20:
                reject_reasons['混合名过长'] += 1
                continue
        else:
            reject_reasons['其他类型'] += 1
            continue
        
        filtered.append((item, lang))
    
    print(f"  通过: {len(filtered)} 条")
    print(f"  淘汰原因:")
    for reason, count in reject_reasons.most_common():
        print(f"    {reason}: {count}")
    
    # === 第二轮：评分 ===
    print(f"\n=== 第二轮：多维度评分 ===")
    scored = []
    
    for item, lang in filtered:
        name = item['name'].strip()
        
        if lang == 'chinese':
            scores = score_chinese(name)
        elif lang == 'english':
            scores = score_english(name)
        elif lang == 'mixed':
            scores = score_mixed(name)
        else:
            continue
        
        scored.append({
            'name': name,
            'language': lang,
            'scores': scores,
            'total': scores['总分'],
            'original': item
        })
    
    # 排序
    scored.sort(key=lambda x: x['total'], reverse=True)
    
    # 统计分数分布
    total_scores = [s['total'] for s in scored]
    if total_scores:
        print(f"  评分范围: {min(total_scores)} - {max(total_scores)}")
        print(f"  平均分: {sum(total_scores) / len(total_scores):.1f}")
        print(f"  中位数: {sorted(total_scores)[len(total_scores)//2]}")
    
    # === 第三轮：取 TOP N ===
    # 确保中英文都有代表
    top_cn = [s for s in scored if s['language'] == 'chinese'][:int(top_n * 0.8)]
    top_en = [s for s in scored if s['language'] == 'english'][:int(top_n * 0.2)]
    top_mixed = [s for s in scored if s['language'] == 'mixed'][:5]  # 最多5个混合
    
    # 合并并重新排序
    combined = top_cn + top_en + top_mixed
    combined.sort(key=lambda x: x['total'], reverse=True)
    
    # 去重
    seen = set()
    unique = []
    for item in combined:
        if item['name'] not in seen:
            seen.add(item['name'])
            unique.append(item)
    
    result = unique[:top_n]
    
    print(f"\n=== 最终结果 ===")
    print(f"  TOP {top_n} 中:")
    lang_counts = Counter(r['language'] for r in result)
    for lang, count in lang_counts.most_common():
        print(f"    {lang}: {count}")
    
    return result

def export_to_excel(results: List[Dict], output_path: str):
    """导出结果到Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TOP50 商标推荐"
    
    # 样式定义
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='微软雅黑', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    name_font = Font(name='微软雅黑', size=14, bold=True, color='1F4E79')
    
    thin_border = Border(
        left=Side(style='thin', color='D6DCE4'),
        right=Side(style='thin', color='D6DCE4'),
        top=Side(style='thin', color='D6DCE4'),
        bottom=Side(style='thin', color='D6DCE4')
    )
    
    gold_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    silver_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    bronze_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    
    # 标题行
    headers = ['排名', '商标名', '语言', '总分', '简洁度\n(25)', '音韵美感\n(20)', 
               '语义品质\n(25)', '品牌感\n(20)', '记忆度\n(10)', '商标图片URL']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 数据行
    for i, item in enumerate(results, 1):
        row = i + 1
        scores = item['scores']
        
        # 排名
        ws.cell(row=row, column=1, value=i).font = data_font
        ws.cell(row=row, column=1).alignment = data_alignment
        
        # 商标名（大号加粗）
        name_cell = ws.cell(row=row, column=2, value=item['name'])
        name_cell.font = name_font
        name_cell.alignment = data_alignment
        
        # 语言
        lang_map = {'chinese': '中文', 'english': '英文', 'mixed': '中英混合'}
        ws.cell(row=row, column=3, value=lang_map.get(item['language'], item['language']))
        ws.cell(row=row, column=3).font = data_font
        ws.cell(row=row, column=3).alignment = data_alignment
        
        # 总分
        ws.cell(row=row, column=4, value=scores['总分']).font = Font(name='微软雅黑', size=12, bold=True)
        ws.cell(row=row, column=4).alignment = data_alignment
        
        # 各维度分数
        ws.cell(row=row, column=5, value=scores['简洁度']).font = data_font
        ws.cell(row=row, column=6, value=scores['音韵美感']).font = data_font
        ws.cell(row=row, column=7, value=scores['语义品质']).font = data_font
        ws.cell(row=row, column=8, value=scores['品牌感']).font = data_font
        ws.cell(row=row, column=9, value=scores['记忆度']).font = data_font
        
        for col in range(5, 10):
            ws.cell(row=row, column=col).alignment = data_alignment
        
        # 图片URL
        img_url = item['original'].get('imgUrl', '')
        ws.cell(row=row, column=10, value=img_url).font = Font(name='微软雅黑', size=9, color='6699CC')
        
        # 行颜色（前3名特殊颜色）
        if i <= 3:
            fill = [gold_fill, silver_fill, bronze_fill][i-1]
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = fill
        
        # 边框
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border
    
    # 列宽
    col_widths = [6, 18, 10, 8, 10, 10, 10, 10, 10, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 行高
    ws.row_dimensions[1].height = 35
    for i in range(2, len(results) + 2):
        ws.row_dimensions[i].height = 28
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"\n✅ 已导出到: {output_path}")

# ============================================================
# 入口
# ============================================================

def main():
    import sys
    
    # 配置
    data_path = '/home/yangyy/projects/biaobiaoku-scraper/all_trademarks_0306_final.json'
    output_dir = '/home/yangyy/projects/biaobiaoku-scraper/antigravity-task'
    output_excel = f'{output_dir}/top50_trademarks.xlsx'
    output_json = f'{output_dir}/top50_trademarks.json'
    top_n = 50
    
    # 读取数据
    print(f"📁 读取数据: {data_path}")
    with open(data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    print(f"   共 {len(data)} 条\n")
    
    # 筛选
    results = filter_trademarks(data, top_n=top_n)
    
    # 显示结果
    print(f"\n{'='*60}")
    print(f"🏆 TOP {len(results)} 商标推荐")
    print(f"{'='*60}\n")
    
    for i, item in enumerate(results, 1):
        s = item['scores']
        medal = '🥇' if i <= 3 else ('🥈' if i <= 10 else ('🥉' if i <= 20 else '  '))
        print(f"{medal} {i:2d}. {item['name']:12s}  [{item['language']:7s}]  "
              f"总分:{s['总分']:3d}  简洁:{s['简洁度']:2d}  音韵:{s['音韵美感']:2d}  "
              f"语义:{s['语义品质']:2d}  品牌:{s['品牌感']:2d}  记忆:{s['记忆度']:2d}")
    
    # 导出
    export_to_excel(results, output_excel)
    
    # 同时保存JSON
    json_results = []
    for i, item in enumerate(results, 1):
        json_results.append({
            'rank': i,
            'name': item['name'],
            'language': item['language'],
            'scores': item['scores'],
            'imgUrl': item['original'].get('imgUrl', ''),
            'category': item['original'].get('category', ''),
        })
    
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(json_results, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON 结果已保存: {output_json}")

if __name__ == '__main__':
    main()
